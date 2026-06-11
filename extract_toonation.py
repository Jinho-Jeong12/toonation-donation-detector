"""
투네이션 후원 감지기 v2
A조건 (화면): 파란 글씨 텍스트 + "~님이 ~캐시" / "영상 후원" 등 패턴 인식
B조건 (음성): 투네이션 효과음 (1000원 / 10000원) 오디오 매칭
A OR B 감지 시 → 전체 방송 화면 스크린샷 저장 (영상별 폴더)
"""

import sys
import csv
import re
import json
import subprocess
import tempfile
import shutil
import zipfile
import urllib.request
from pathlib import Path

import warnings
warnings.filterwarnings("ignore", message="'pin_memory' argument is set as true")


# ── 패키지 자동 설치 ──────────────────────────────────────────────────────────

def _ensure_packages():
    required = ["scipy", "openpyxl", "easyocr", "opencv-python"]
    import importlib
    pkg_map = {
        "scipy": "scipy",
        "openpyxl": "openpyxl",
        "easyocr": "easyocr",
        "opencv-python": "cv2",
    }
    for pip_name, import_name in pkg_map.items():
        try:
            importlib.import_module(import_name)
        except ImportError:
            print(f"[자동 설치] {pip_name} 설치 중...")
            subprocess.run(
                [sys.executable, "-m", "pip", "install", pip_name],
                check=True,
            )
            print(f"[자동 설치] {pip_name} 완료\n")

_ensure_packages()

import cv2
import numpy as np


# ── ffmpeg 자동 설치 ──────────────────────────────────────────────────────────

FFMPEG_DIR = Path(__file__).parent / "ffmpeg_bin"
FFMPEG_EXE = FFMPEG_DIR / "ffmpeg.exe"
# BtbN 빌드 (GitHub Releases) — Windows 64bit GPL essentials
FFMPEG_URL = "https://github.com/BtbN/FFmpeg-Builds/releases/download/latest/ffmpeg-master-latest-win64-gpl-shared.zip"


def _find_ffmpeg_in_path() -> str | None:
    return shutil.which("ffmpeg")


def ensure_ffmpeg() -> str:
    """
    ffmpeg.exe 경로 반환.
    시스템 PATH에 있으면 그것 사용, 없으면 스크립트 옆 ffmpeg_bin/ 에 자동 다운로드.
    """
    # 1) 시스템 PATH 확인
    sys_ffmpeg = _find_ffmpeg_in_path()
    if sys_ffmpeg:
        return sys_ffmpeg

    # 2) 로컬 캐시 확인
    if FFMPEG_EXE.exists():
        return str(FFMPEG_EXE)

    # 3) 자동 다운로드
    print("=" * 60)
    print("ffmpeg가 설치돼 있지 않습니다. 자동으로 다운로드합니다.")
    print(f"저장 위치: {FFMPEG_DIR}")
    print("(한 번만 다운로드되며 이후엔 재사용됩니다)")
    print("=" * 60)

    FFMPEG_DIR.mkdir(exist_ok=True)
    zip_path = FFMPEG_DIR / "ffmpeg.zip"

    try:
        print("다운로드 중... (수십 MB, 잠시 기다려 주세요)")

        def _progress(block_num, block_size, total_size):
            if total_size > 0:
                pct = min(100, block_num * block_size * 100 // total_size)
                print(f"\r  {pct}%", end="", flush=True)

        urllib.request.urlretrieve(FFMPEG_URL, zip_path, reporthook=_progress)
        print("\n다운로드 완료. 압축 해제 중...")

        with zipfile.ZipFile(zip_path, "r") as zf:
            # zip 내부에서 ffmpeg.exe 찾아 ffmpeg_bin/ 에 바로 꺼냄
            for member in zf.namelist():
                if member.endswith("/bin/ffmpeg.exe"):
                    with zf.open(member) as src, open(FFMPEG_EXE, "wb") as dst:
                        dst.write(src.read())
                    break
            else:
                # 혹시 경로 구조가 다르면 그냥 전체 추출 후 탐색
                zf.extractall(FFMPEG_DIR / "_tmp")
                found = list((FFMPEG_DIR / "_tmp").rglob("ffmpeg.exe"))
                if found:
                    shutil.copy(found[0], FFMPEG_EXE)
                shutil.rmtree(FFMPEG_DIR / "_tmp", ignore_errors=True)

        zip_path.unlink(missing_ok=True)

        if FFMPEG_EXE.exists():
            print(f"ffmpeg 준비 완료: {FFMPEG_EXE}\n")
            return str(FFMPEG_EXE)
        else:
            raise FileNotFoundError("ffmpeg.exe를 찾지 못했습니다.")

    except Exception as e:
        zip_path.unlink(missing_ok=True)
        print(f"\n[오류] ffmpeg 자동 다운로드 실패: {e}")
        print("수동으로 https://ffmpeg.org/download.html 에서 설치하거나")
        print("  winget install ffmpeg  를 실행해 주세요.")
        print()
        print("┌─────────────────────────────────────────────────────────┐")
        print("│  ⚠  경고: ffmpeg 없이 계속 진행합니다                  │")
        print("│                                                         │")
        print("│  음성 감지(B조건)를 사용할 수 없습니다.                │")
        print("│  화면 텍스트 인식(A조건)만으로 동작하므로              │")
        print("│  일부 후원을 놓칠 수 있습니다.                         │")
        print("│                                                         │")
        print("│  계속하려면 엔터, 종료하려면 Ctrl+C 를 누르세요.      │")
        print("└─────────────────────────────────────────────────────────┘")
        input()
        return ""


# ── 설정 ──────────────────────────────────────────────────────────────────────
SCAN_INTERVAL_SEC = 3.0    # 프레임 스캔 간격 (초)

ROI_X_START = 0.20
ROI_X_END   = 0.80
ROI_Y_START = 0.0
ROI_Y_END   = 1.0

BLUE_HSV_LOWER = np.array([80, 80, 80])   # 느슨하게
BLUE_HSV_UPPER = np.array([140, 255, 255])
BLUE_PIXEL_THRESHOLD = 40                  # 느슨하게 (기존 80 → 40)

DEDUPE_SEC = 8             # 중복 감지 방지 간격

OCR_MAX_WIDTH = 640

# 음성 감지
AUDIO_SR = 8000            # 8kHz 모노로 추출 (속도/메모리 절약)
AUDIO_THRESHOLD = 0.15     # 느슨하게 (낮을수록 더 많이 잡음)
AUDIO_MIN_GAP_SEC = 8      # 음성 피크 최소 간격

# 효과음 파일 경로 (스크립트와 같은 폴더)
SCRIPT_DIR = Path(__file__).parent
SFX_PATHS = [
    SCRIPT_DIR / "투네이션 1,000원 효과음.mp3",
    SCRIPT_DIR / "투네이션 10,000원 후원 효과음.mp3",
]
# ─────────────────────────────────────────────────────────────────────────────


def sec_to_timestamp(sec: float) -> str:
    total = int(sec)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def timestamp_to_sec(ts: str) -> float:
    parts = [int(p) for p in str(ts).strip().split(":")]
    if len(parts) == 3:
        return parts[0] * 3600 + parts[1] * 60 + parts[2]
    return parts[0] * 60 + parts[1]


# ── 텍스트 파싱 (A조건) ───────────────────────────────────────────────────────

def parse_toonation(text: str) -> dict | None:
    """텍스트에서 투네이션 후원 감지. 느슨하게 여러 패턴 시도."""

    # 패턴 1: "닉네임 님이 N캐시" (기존)
    m = re.search(r"(.+?)\s*님이?\s*([\d,]+)\s*캐시", text, re.IGNORECASE)
    if m:
        donor = m.group(1).strip()
        try:
            amount = int(m.group(2).replace(",", ""))
        except ValueError:
            amount = 0
        return {"닉네임": donor, "금액(캐시)": amount, "원본텍스트": text.strip(), "출처": "텍스트-패턴1"}

    # 패턴 2: "영상 후원"
    if re.search(r"영상\s*후원", text, re.IGNORECASE):
        am = re.search(r"([\d,]+)\s*캐시", text)
        amount = int(am.group(1).replace(",", "")) if am else 0
        nm = re.search(r"(.+?)\s*님", text)
        donor = nm.group(1).strip() if nm else "영상후원"
        return {"닉네임": donor, "금액(캐시)": amount, "원본텍스트": text.strip(), "출처": "텍스트-영상후원"}

    # 패턴 3: "후원" + 캐시 숫자 (느슨)
    if re.search(r"후원", text, re.IGNORECASE):
        am = re.search(r"([\d,]+)\s*캐시", text)
        if am:
            try:
                amount = int(am.group(1).replace(",", ""))
            except ValueError:
                amount = 0
            nm = re.search(r"(.+?)\s*님", text)
            donor = nm.group(1).strip() if nm else "후원감지"
            return {"닉네임": donor, "금액(캐시)": amount, "원본텍스트": text.strip(), "출처": "텍스트-느슨"}

    # 패턴 4: "캐시" 단독 (매우 느슨 - 파란 화면 + 캐시 단어만 있어도)
    if re.search(r"\d[\d,]*\s*캐시", text, re.IGNORECASE):
        am = re.search(r"([\d,]+)\s*캐시", text)
        amount = int(am.group(1).replace(",", "")) if am else 0
        return {"닉네임": "?", "금액(캐시)": amount, "원본텍스트": text.strip(), "출처": "텍스트-캐시만"}

    return None


# ── 음성 감지 (B조건) ─────────────────────────────────────────────────────────

def _ffmpeg_to_wav(src_path: Path, out_path: Path, sr: int = AUDIO_SR,
                   ffmpeg_exe: str = "") -> bool:
    """ffmpeg로 오디오를 모노 wav로 변환. 성공 여부 반환."""
    exe = ffmpeg_exe or _find_ffmpeg_in_path() or str(FFMPEG_EXE)
    if not exe:
        return False
    try:
        result = subprocess.run(
            [
                exe, "-y", "-i", str(src_path),
                "-vn", "-ar", str(sr), "-ac", "1",
                "-f", "wav", str(out_path),
            ],
            capture_output=True, timeout=600,
        )
        return result.returncode == 0 and out_path.exists()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _load_wav(path: Path) -> tuple[np.ndarray, int] | tuple[None, None]:
    """wav 파일 로드. (float32 배열, sr) 또는 (None, None)."""
    try:
        from scipy.io import wavfile
        sr, data = wavfile.read(str(path))
        if data.ndim > 1:
            data = data.mean(axis=1)
        data = data.astype(np.float32)
        max_val = np.abs(data).max()
        if max_val > 0:
            data /= max_val
        return data, sr
    except Exception as e:
        print(f"  [음성] wav 읽기 실패: {e}")
        return None, None


def detect_sound_events(video_path: Path) -> list[float]:
    """
    영상에서 투네이션 효과음 감지.
    감지된 시간(초) 목록 반환. ffmpeg 없으면 빈 리스트.
    """
    available_sfx = [p for p in SFX_PATHS if p.exists()]
    if not available_sfx:
        print("  [음성] 효과음 파일 없음 — 음성 감지 건너뜀")
        return []

    ffmpeg_exe = ensure_ffmpeg()
    if not ffmpeg_exe:
        print("  [음성] ffmpeg 없음 — 음성 감지 건너뜀")
        return []

    print("  [음성] 오디오 추출 중 (ffmpeg)...")
    with tempfile.TemporaryDirectory() as tmpdir:
        video_wav = Path(tmpdir) / "video_audio.wav"
        if not _ffmpeg_to_wav(video_path, video_wav, ffmpeg_exe=ffmpeg_exe):
            print("  [음성] ffmpeg 변환 실패 — 음성 감지 건너뜀")
            return []

        video_audio, sr = _load_wav(video_wav)
        if video_audio is None:
            return []

        print(f"  [음성] 오디오 로드 완료 ({len(video_audio)/sr:.1f}초, {sr}Hz)")

        all_detections: list[float] = []

        for sfx_path in available_sfx:
            sfx_wav = Path(tmpdir) / f"{sfx_path.stem}.wav"
            if not _ffmpeg_to_wav(sfx_path, sfx_wav, sr, ffmpeg_exe=ffmpeg_exe):
                continue
            template, _ = _load_wav(sfx_wav)
            if template is None or len(template) == 0:
                continue

            print(f"  [음성] '{sfx_path.name}' 매칭 중 ({len(template)/sr:.2f}초 템플릿)...")

            try:
                from scipy.signal import correlate, find_peaks

                # normalized cross-correlation
                corr = correlate(video_audio, template, mode="valid")

                template_energy = np.sum(template ** 2)
                tlen = len(template)
                # running window energy
                sq = video_audio ** 2
                cum = np.concatenate([[0.0], np.cumsum(sq)])
                win_energy = cum[tlen:] - cum[:len(cum) - tlen]
                denom = np.sqrt(template_energy * win_energy + 1e-10)
                corr_norm = corr / denom
                corr_norm = np.clip(corr_norm, -1.0, 1.0)

                min_dist = max(1, int(AUDIO_MIN_GAP_SEC * sr))
                peaks, _ = find_peaks(corr_norm, height=AUDIO_THRESHOLD, distance=min_dist)

                times = [float(p / sr) for p in peaks]
                print(f"    → {len(times)}건 감지 (threshold={AUDIO_THRESHOLD})")
                all_detections.extend(times)

            except ImportError:
                print("  [음성] scipy 미설치 — 음성 감지 건너뜀")
                return []

        # 정렬 후 중복 제거
        all_detections.sort()
        deduped: list[float] = []
        last_t = -9999.0
        for t in all_detections:
            if t - last_t >= AUDIO_MIN_GAP_SEC:
                deduped.append(t)
                last_t = t

        print(f"  [음성] 최종 감지: {len(deduped)}건\n")
        return deduped


# ── 화면 처리 헬퍼 ────────────────────────────────────────────────────────────

def crop_roi(frame: np.ndarray) -> np.ndarray:
    h, w = frame.shape[:2]
    return frame[
        int(h * ROI_Y_START):int(h * ROI_Y_END),
        int(w * ROI_X_START):int(w * ROI_X_END),
    ]


def has_blue_text(roi_bgr: np.ndarray) -> bool:
    hsv = cv2.cvtColor(roi_bgr, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, BLUE_HSV_LOWER, BLUE_HSV_UPPER)
    return int(np.count_nonzero(mask)) >= BLUE_PIXEL_THRESHOLD


def save_screenshot(frame: np.ndarray, screenshot_dir: Path,
                    current_sec: float, info: dict | None, source: str) -> str:
    """전체 프레임 스크린샷 저장. 파일명 반환."""
    safe_ts = sec_to_timestamp(current_sec).replace(":", "-")
    if source == "audio":
        filename = f"{safe_ts}_음성감지.png"
    else:
        donor = str(info.get("닉네임", "?"))[:20]
        amount = info.get("금액(캐시)", "?")
        src_tag = info.get("출처", "텍스트")
        safe_donor = re.sub(r'[\\/:*?"<>|]', "_", donor)
        filename = f"{safe_ts}_{safe_donor}_{amount}캐시.png"
    out_path = screenshot_dir / filename
    # cv2.imwrite는 Windows 한글 경로에서 실패할 수 있으므로 imencode 우회
    ret, buf = cv2.imencode(".png", frame)
    if ret:
        out_path.write_bytes(buf.tobytes())
    return filename


# ── Excel / CSV ────────────────────────────────────────────────────────────────

def _row_style(ws, row_num: int, is_even: bool):
    from openpyxl.styles import PatternFill, Alignment, Border, Side
    fill = PatternFill("solid", fgColor="E3F2FD" if is_even else "FFFFFF")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")


def open_or_create_xlsx(xlsx_path: Path) -> tuple:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    existing_results = []
    resume_sec = -1.0

    if xlsx_path.exists():
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            if ws.max_row > 1 and str(ws.cell(ws.max_row, 1).value).strip() == "합계":
                ws.delete_rows(ws.max_row)
            for row in ws.iter_rows(min_row=2, values_only=True):
                ts = row[0]
                if not ts:
                    continue
                try:
                    sec = timestamp_to_sec(ts)
                except Exception:
                    continue
                existing_results.append({
                    "시간": str(ts).strip(),
                    "닉네임": row[1],
                    "금액(캐시)": row[2],
                    "원본텍스트": row[3] or "",
                    "출처": row[4] if len(row) > 4 else "",
                })
                resume_sec = max(resume_sec, sec)
            print(f"  ↩ 기존 엑셀 발견 — {len(existing_results)}건, {sec_to_timestamp(resume_sec)} 이후부터 재개\n")
            return wb, ws, existing_results, resume_sec
        except Exception as e:
            print(f"  [경고] 기존 엑셀 읽기 실패: {e} — 새로 시작합니다.\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "투네이션 후원 목록"
    ws.append(["시간", "닉네임", "금액 (캐시)", "원본 텍스트", "출처"])
    header_fill = PatternFill("solid", fgColor="1565C0")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 15
    ws.row_dimensions[1].height = 22
    wb.save(xlsx_path)
    return wb, ws, [], -1.0


def append_to_xlsx(wb, ws, xlsx_path: Path, entry: dict, row_idx: int):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    ws.append([
        entry["시간"], entry["닉네임"], entry["금액(캐시)"],
        entry.get("원본텍스트", ""), entry.get("출처", ""),
    ])
    _row_style(ws, ws.max_row, row_idx % 2 == 0)
    wb.save(xlsx_path)


def finalize_xlsx(wb, ws, xlsx_path: Path):
    from openpyxl.styles import Font, PatternFill
    if ws.max_row < 2:
        return
    ws.append(["합계", "", f"=SUM(C2:C{ws.max_row})", "", ""])
    total_row = ws.max_row
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"].font = Font(bold=True)
    ws[f"C{total_row}"].fill = PatternFill("solid", fgColor="FFF9C4")
    wb.save(xlsx_path)


# ── 진행 파일 ─────────────────────────────────────────────────────────────────

def progress_path(video_path: Path) -> Path:
    return video_path.parent / f"{video_path.stem}_progress.json"


def load_progress(video_path: Path) -> int | None:
    p = progress_path(video_path)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))["frame_idx"]
        except Exception:
            pass
    return None


def save_progress(video_path: Path, frame_idx: int):
    progress_path(video_path).write_text(
        json.dumps({"frame_idx": frame_idx}), encoding="utf-8"
    )


# ── 메인 처리 ─────────────────────────────────────────────────────────────────

def process_video(video_path: Path, output_dir: Path):
    import easyocr

    # 스크린샷 폴더
    screenshot_dir = output_dir / f"{video_path.stem}_스크린샷"
    screenshot_dir.mkdir(exist_ok=True)
    print(f"스크린샷 저장 폴더: {screenshot_dir}\n")

    xlsx_path = output_dir / f"{video_path.stem}_투네이션.xlsx"
    wb, ws, existing_results, resume_sec = open_or_create_xlsx(xlsx_path)
    total_found = len(existing_results)

    # B조건: 음성 감지 — 별도 스레드로 A조건과 동시에 실행
    import threading
    audio_detections: list[float] = []
    audio_done = threading.Event()

    def _run_audio():
        results = detect_sound_events(video_path)
        audio_detections.extend(results)
        audio_done.set()

    audio_thread = threading.Thread(target=_run_audio, daemon=True)
    audio_thread.start()
    print("=== B조건(음성) 백그라운드 시작 / A조건(화면) 동시 시작 ===\n")

    print("EasyOCR 한국어 모델 로드 중...")
    reader = easyocr.Reader(["ko", "en"], gpu=False)
    print("모델 로드 완료\n")

    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        print(f"[오류] 영상 파일을 열 수 없습니다: {video_path}")
        return total_found

    fps = cap.get(cv2.CAP_PROP_FPS) or 30
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    total_sec = total_frames / fps
    frame_step = max(1, int(fps * SCAN_INTERVAL_SEC))

    saved_frame = load_progress(video_path)
    if saved_frame is not None:
        start_frame = (saved_frame // frame_step) * frame_step
        print(f"  ↩ 진행 파일 발견 — {sec_to_timestamp(start_frame / fps)} 부터 재개\n")
    elif resume_sec >= 0:
        start_frame = (int(max(0.0, resume_sec - DEDUPE_SEC) * fps) // frame_step) * frame_step
    else:
        start_frame = 0

    last_saved_sec = resume_sec if resume_sec >= 0 else -(DEDUPE_SEC + 1)

    print(f"영상: {total_sec/3600:.1f}시간  |  {fps:.1f}fps  |  {total_frames:,}프레임")
    print(f"스캔: {SCAN_INTERVAL_SEC}초 간격  |  시작: {sec_to_timestamp(start_frame / fps)}\n")

    audio_idx = 0  # audio_detections 포인터

    cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)
    current_frame = start_frame

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        current_sec = current_frame / fps

        if current_frame % (frame_step * 60) == 0:
            pct = current_sec / total_sec * 100
            print(f"  [{pct:5.1f}%] {sec_to_timestamp(current_sec)} ... 발견 {total_found}건")
            save_progress(video_path, current_frame)

        # ── B조건 체크 (음성 결과는 스레드에서 실시간으로 채워짐) ──────────
        sorted_audio = sorted(audio_detections)

        while audio_idx < len(sorted_audio) and sorted_audio[audio_idx] < current_sec - SCAN_INTERVAL_SEC:
            missed_sec = sorted_audio[audio_idx]
            if missed_sec - last_saved_sec >= DEDUPE_SEC:
                fname = save_screenshot(frame, screenshot_dir, missed_sec, None, "audio")
                entry = {
                    "시간": sec_to_timestamp(missed_sec),
                    "닉네임": "음성감지",
                    "금액(캐시)": 0,
                    "원본텍스트": "효과음 감지",
                    "출처": "음성-B조건",
                }
                append_to_xlsx(wb, ws, xlsx_path, entry, total_found)
                total_found += 1
                last_saved_sec = missed_sec
                print(f"  ♪ [{sec_to_timestamp(missed_sec)}] 음성 감지 → {fname}")
            audio_idx += 1

        b_triggered = False
        if audio_idx < len(sorted_audio):
            if abs(current_sec - sorted_audio[audio_idx]) <= SCAN_INTERVAL_SEC:
                b_triggered = True

        # ── A조건 체크 ───────────────────────────────────────────────────────
        a_triggered = False
        a_info = None
        roi = crop_roi(frame)
        if has_blue_text(roi):
            h_roi, w_roi = roi.shape[:2]
            if w_roi > OCR_MAX_WIDTH:
                roi_ocr = cv2.resize(roi, (OCR_MAX_WIDTH, int(h_roi * OCR_MAX_WIDTH / w_roi)))
            else:
                roi_ocr = roi
            texts = reader.readtext(roi_ocr, detail=0, paragraph=True)
            full_text = " ".join(texts)
            parsed = parse_toonation(full_text)
            if parsed:
                a_triggered = True
                a_info = parsed

        # ── OR 조건으로 스크린샷 저장 ────────────────────────────────────────
        if (a_triggered or b_triggered) and (current_sec - last_saved_sec >= DEDUPE_SEC):
            if a_triggered:
                source = "텍스트-A조건"
                info = a_info
                fname = save_screenshot(frame, screenshot_dir, current_sec, a_info, "text")
            else:
                source = "음성-B조건"
                info = {"닉네임": "음성감지", "금액(캐시)": 0, "출처": "음성-B조건"}
                fname = save_screenshot(frame, screenshot_dir, current_sec, None, "audio")
                # 음성 포인터 전진
                if audio_idx < len(sorted_audio) and abs(current_sec - sorted_audio[audio_idx]) <= SCAN_INTERVAL_SEC:
                    audio_idx += 1

            entry = {
                "시간": sec_to_timestamp(current_sec),
                "닉네임": info.get("닉네임", "?"),
                "금액(캐시)": info.get("금액(캐시)", 0),
                "원본텍스트": info.get("원본텍스트", ""),
                "출처": source,
            }
            append_to_xlsx(wb, ws, xlsx_path, entry, total_found)
            total_found += 1
            last_saved_sec = current_sec
            icon = "★" if a_triggered else "♪"
            print(f"  {icon} [{entry['시간']}] {entry['닉네임']}  {entry['금액(캐시)']}캐시  ({source})  → {fname}")

        current_frame += frame_step
        for _ in range(frame_step - 1):
            if not cap.grab():
                break

    cap.release()

    # 음성 스레드가 아직 돌고 있으면 완료 대기 후 남은 감지 결과 처리
    if not audio_done.is_set():
        print("\n  [음성] 분석 완료 대기 중...")
        audio_done.wait()

    sorted_audio = sorted(audio_detections)
    while audio_idx < len(sorted_audio):
        missed_sec = sorted_audio[audio_idx]
        if missed_sec - last_saved_sec >= DEDUPE_SEC:
            print(f"  ♪ [{sec_to_timestamp(missed_sec)}] 음성 감지 (후처리) — 스크린샷 없음")
            entry = {
                "시간": sec_to_timestamp(missed_sec),
                "닉네임": "음성감지",
                "금액(캐시)": 0,
                "원본텍스트": "효과음 감지 (영상 스캔 후 처리)",
                "출처": "음성-B조건",
            }
            append_to_xlsx(wb, ws, xlsx_path, entry, total_found)
            total_found += 1
            last_saved_sec = missed_sec
        audio_idx += 1

    finalize_xlsx(wb, ws, xlsx_path)
    progress_path(video_path).unlink(missing_ok=True)
    return total_found


def save_csv(xlsx_path: Path, csv_path: Path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        fields = ["시간", "닉네임", "금액(캐시)", "원본텍스트", "출처"]
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for row in ws.iter_rows(min_row=2, values_only=True):
                ts = row[0]
                if not ts or str(ts).strip() == "합계":
                    continue
                w.writerow({
                    "시간": row[0], "닉네임": row[1], "금액(캐시)": row[2],
                    "원본텍스트": row[3] or "", "출처": row[4] or "",
                })
    except Exception as e:
        print(f"CSV 저장 실패: {e}")


def pick_video_file() -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="투네이션 추출할 영상 파일 선택",
            filetypes=[("영상 파일", "*.mp4 *.mkv *.webm *.avi *.mov"), ("모든 파일", "*.*")],
        )
        root.destroy()
        return Path(path) if path else None
    except Exception:
        return None


def main():
    if len(sys.argv) >= 2:
        video_path = Path(sys.argv[1])
    else:
        print("=== 투네이션 후원 감지기 v2 ===")
        print("영상 파일 선택 창을 엽니다...\n")
        video_path = pick_video_file()
        if not video_path:
            print("파일을 선택하지 않았습니다. 종료합니다.")
            input("\n엔터를 누르면 창이 닫힙니다.")
            sys.exit(0)

    if not video_path.exists():
        print(f"[오류] 파일 없음: {video_path}")
        input("\n엔터를 누르면 창이 닫힙니다.")
        sys.exit(1)

    output_dir = Path(__file__).parent
    xlsx_path = output_dir / f"{video_path.stem}_투네이션.xlsx"
    screenshot_dir = output_dir / f"{video_path.stem}_스크린샷"

    print(f"=== 투네이션 후원 감지기 v2 ===")
    print(f"영상:       {video_path}")
    print(f"Excel:      {xlsx_path}")
    print(f"스크린샷:   {screenshot_dir}")
    print(f"A조건 (화면) OR B조건 (음성) — 둘 중 하나라도 감지 시 스크린샷\n")

    total = process_video(video_path, output_dir)

    xlsx_path = output_dir / f"{video_path.stem}_투네이션.xlsx"
    csv_path  = output_dir / f"{video_path.stem}_투네이션.csv"
    if total > 0:
        save_csv(xlsx_path, csv_path)
        print(f"\n[스크린샷] {screenshot_dir}  ({total}건)")
        print(f"[Excel]    {xlsx_path}")
        print(f"[CSV]      {csv_path}")
    else:
        print("\n투네이션 후원을 감지하지 못했습니다.")
        print("  → ffmpeg 설치 확인, 효과음 파일 위치 확인, AUDIO_THRESHOLD 조정 등 시도해보세요.")

    print("\n완료!")
    input("\n엔터를 누르면 창이 닫힙니다.")


if __name__ == "__main__":
    main()
